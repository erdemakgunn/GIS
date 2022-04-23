import veri_cekme
import folium
from folium.plugins import MarkerCluster
from openpyxl import Workbook, load_workbook
from folium.plugins import MousePosition
from folium.plugins import FloatImage
from folium.plugins import Draw

m = folium.Map(location=[41, 29], zoom_start=12, tiles="Stamen Terrain")

#Fare koordinatı göstermek için
MousePosition().add_to(m)

#Görsel ikon ekleme
image_file = 'src//saglik_logo.png'
image_file2 = 'src//itfaiye_logo.png'
FloatImage(image_file, bottom=0, left=0).add_to(m)
FloatImage(image_file2, bottom=0, left=5).add_to(m)


marker_cluster = MarkerCluster().add_to(m)
marker_cluster2 = MarkerCluster().add_to(m)

marker_cluster.layer_name = 'itfaiye'
marker_cluster2.layer_name = 'saglik'

image_file = 'src//saglik_logo.png'
image_file2 = 'src//itfaiye_logo.png'
# FloatImage(image_file, bottom=0, left=0).add_to(m)
# FloatImage(image_file2, bottom=0, left=5).add_to(m)

for index, row in veri_cekme.istasyon_kordinat_isim.iterrows():
    coordinat = veri_cekme.istasyon_kordinat_isim.loc[index, 'Koordinat']
    ad = veri_cekme.istasyon_kordinat_isim.loc[index, 'ITFAIYE_BIRIM_ADI']

    icon_url = "src//itfaiye_logo.png"
    icon = folium.features.CustomIcon(icon_url,
                                      icon_size=(54, 54))
    folium.Marker(
        coordinat.split(','), popup=ad, icon=icon,
    ).add_to(marker_cluster)

wb = load_workbook("13064,1basamakmevcutkoordinatxlsx.xlsx")
ws = wb.active

for satir in range(7151, 8328):
    for sutun in range(9, 10):
        xcoordinat = ws.cell(satir, sutun).value
        # print(xcoordinat)

    for sutun in range(10, 11):
        ycoordinat = ws.cell(satir, sutun).value

    for sutun in range(5, 6):
        ad = []
        ad = ws.cell(satir, sutun).value
        # print(ad)

    icon_url = "src//saglik_logo.png"
    icon = folium.features.CustomIcon(icon_url,
                                      icon_size=(54, 55))
    folium.Marker(
        location=[xcoordinat, ycoordinat], popup=ad, icon=icon
    ).add_to(marker_cluster2)


draw = Draw(
    draw_options={
        'polyline': True,
        'rectangle': True,
        'polygon': True,
        'circle': True,
        'marker': True,
        'circlemarker': True},
    edit_options={'edit': True})
m.add_child(draw)

folium.LayerControl().add_to(m)

m.save("staj.html")
